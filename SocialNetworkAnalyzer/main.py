import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

import config as cf
import utils
import ok
import vk
import tg


def fill_cells(sheet, stat, indexes):
    """
    fills Excel cells with special colours. Red - for min-values, green - for max-values
    :param sheet: sheet of Excel file
    :param stat: stat-table retrieved from network APIs
    :param indexes: list of column numbers
    :return: None
    """
    ext_list = list(zip(*stat))
    ext_dict = {}
    for i in indexes:
        ext_dict[i] = (max(ext_list[i - 1]), min(ext_list[i - 1]))

    i = sheet.max_row + 1
    for j in range(len(ext_list)):
        cell = sheet.cell(i, j + 1)
        cell.fill = PatternFill('solid', fgColor="808080")

    for rec_ in stat:
        i += 1
        sheet.append(rec_)
        for j in ext_dict.keys():
            cell = sheet.cell(row=i, column=j)
            if ext_dict[j][0] == ext_dict[j][1]:
                continue
            if cell.value == ext_dict[j][0] and ext_list[j - 1].count(ext_dict[j][0]) == 1:
                cell.fill = PatternFill('solid', fgColor="00FF00")
            elif cell.value == ext_dict[j][1] and ext_list[j - 1].count(ext_dict[j][1]) == 1:
                cell.fill = PatternFill('solid', fgColor="FF0000")


def edit_sheet(sheet_name, column_names):
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name, 0)
    sheet.delete_rows(1, 100)
    sheet.append(column_names)
    return sheet


def download_all_stats():
    if is_new:
        ok_sheet = edit_sheet(cf.OK_SHEET_NAME, cf.OK_COLUMNS)
        vk_sheet = edit_sheet(cf.VK_SHEET_NAME, cf.VK_COLUMNS)
        tg_sheet = edit_sheet(cf.TG_SHEET_NAME, cf.TG_COLUMNS)
    else:
        ok_sheet = wb[cf.OK_SHEET_NAME]
        vk_sheet = wb[cf.VK_SHEET_NAME]
        tg_sheet = wb[cf.TG_SHEET_NAME]

    ok_stat = ok_analyser.get_stat_by_last_week()
    fill_cells(ok_sheet, ok_stat, [4, 5, 6, 7])

    vk_stat = vk_analyser.get_stat_by_last_week()
    fill_cells(vk_sheet, vk_stat, [4, 5, 6, 7, 8, 9])

    tg_stat = tg_analyser.get_stat_by_last_week()
    fill_cells(tg_sheet, tg_stat, [4, 5, 6, 7, 8, 9, 10])


if os.path.isfile('Analytics.xlsx'):
    wb = openpyxl.load_workbook('Analytics.xlsx')
    is_new = False
else:
    wb = openpyxl.Workbook()
    is_new = True

ok_analyser = ok.OKAnalyser()
vk_analyser = vk.VKAnalyser()
tg_analyser = tg.TGAnalyser()

env = utils.show_menu_and_get_choice()
choice = env[1]
env = env[0]

if choice == len(env) + 1:
    for num, file in env.items():
        if num == len(env) + 1:
            break
        utils.init_profile_settings(file)
        download_all_stats()
else:
    utils.init_profile_settings(env[choice])
    download_all_stats()

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

wb.save('Analytics.xlsx')